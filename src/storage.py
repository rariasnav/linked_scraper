import os
from openpyxl import Workbook, load_workbook
from typing import List, Dict


def save_to_excel(rows: List[Dict[str, str]], filename: str):
    """Append rows to an Excel file. If file doesn't exist, create it with headers."""
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
        # remove the all rows to avoid the duplicated.
        ws.delete_rows(1, ws.max_row)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Title", "Company", "Location",
                  "Date", "Description", "Link", "E_Link"])

    for r in rows:
        ws.append([
            r.get("Title", ""),
            r.get("Company", ""),
            r.get("Location", ""),
            r.get("Date", ""),
            r.get("Description", ""),
            r.get("Link", ""),
            r.get("E_Link", "")
        ])

    wb.save(filename)


def load_seen_links(filename: str) -> set:
    """Return a set of links already saved to avoid duplicates."""
    if not os.path.exists(filename):
        return set()

    wb = load_workbook(filename, read_only=True)
    ws = wb.active
    seen = set()

    for row in ws.iter_rows(min_row=2):
        link_cell = row[5]
        if link_cell.value:
            seen.add(link_cell.value)

    return seen
